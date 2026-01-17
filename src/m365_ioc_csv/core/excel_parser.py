"""Excel File Parser Module with Multi-Sheet Support.

Handles parsing Excel (.xlsx, .xls) files with:
- Multi-sheet support
- Sheet enumeration
- Row extraction
- Data-only mode (no formulas, formatting)
- Graceful degradation if openpyxl not installed

Dependencies:
    openpyxl>=3.1.0 (optional)
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, List, Dict, Any

from m365_ioc_csv.utils.error_handler import CSVError, FileError
from m365_ioc_csv.utils.error_handler import get_error_handler
from m365_ioc_csv.utils.logger import get_logger

logger = get_logger(__name__)


@dataclass
class ExcelSheetInfo:
    """Information about a single sheet in Excel workbook."""

    def __init__(
        self,
        name: str,
        index: int,
        row_count: int,
        ioc_count: int
    ):
        """
        Initialize sheet information.

        Args:
            name: Sheet name
            index: Sheet index (0-based)
            row_count: Total number of rows
            ioc_count: Estimated IoC count
        """
        self.name = name
        self.index = index
        self.row_count = row_count
        self.ioc_count = ioc_count

    def __repr__(self) -> str:
        return f"ExcelSheetInfo(name='{self.name}', index={self.index}, rows={self.row_count}, iocs={self.ioc_count})"


@dataclass
class ExcelParseResult:
    """
    Result of parsing an Excel file.

    Attributes:
        file_path: Path to the Excel file
        sheets: List of ExcelSheetInfo
        selected_sheets: List of selected sheet names
        total_rows: Total rows across all selected sheets
        total_iocs: Total IoCs detected across all sheets
        encoding: Always 'utf-8' for Excel
        delimiter: Always 'csv' for consistency with existing code
        sheet_results: Dict mapping sheet names to their data
    """

    file_path: Path
    sheets: List[ExcelSheetInfo] = field(default_factory=list)
    selected_sheets: List[str] = field(default_factory=list)
    total_rows: int = 0
    total_iocs: int = 0
    encoding: str = "utf-8"
    delimiter: str = ","
    sheet_results: Dict[str, List[str]] = field(default_factory=dict)


class ExcelParser:
    """
    Parse Excel (.xlsx, .xls) files with multi-sheet support.

    Features:
    - Sheet enumeration with metadata
    - Sheet selection support
    - Row extraction with data-only mode
    - IoC count estimation per sheet
    - Graceful degradation if openpyxl not installed

    Example:
        parser = ExcelParser()
        sheets = parser.get_sheets(excel_file_path)
        result = parser.parse_sheets(excel_file_path, ['Sheet1', 'Sheet2'])
        for sheet_name, values in result.sheet_results.items():
            print(f"{sheet_name}: {len(values)} values extracted")
    """

    def __init__(self) -> None:
        """Initialize the Excel parser."""
        self.error_handler = get_error_handler()
        logger.debug("ExcelParser initialized")

    # Comment patterns to filter out (matching CSVParser and IoCDetector)
    COMMENT_PATTERNS = ("#", "//", ";", "--")

    # Common non-IoC values to filter out
    NON_IOC_PATTERNS = (
        "N/A", "NA", "n/a", "na", "NULL", "null", "NONE", "none",
        "UNKNOWN", "unknown", "NOT APPLICABLE", "not applicable",
        "-", "—", "–",  # Various dash characters
    )

    def get_sheets(self, file_path: Path) -> List[ExcelSheetInfo]:
        """
        Get information about all sheets in Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            List of ExcelSheetInfo objects

        Raises:
            FileError: If file cannot be read
            CSVError: If Excel file cannot be parsed
        """
        self.error_handler.validate_file(file_path)

        try:
            import openpyxl
        except ImportError:
            raise CSVError(
                "Excel support requires openpyxl. "
                "Install with: pip install openpyxl"
            )

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            raise CSVError(f"Failed to load Excel file: {e}")

        sheets = []

        # Scan all sheets to get metadata
        for sheet in wb.worksheets:
            row_count = 0
            ioc_count = 0

            # Count rows and estimate IoCs (limit to first 100 rows for speed)
            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if idx >= 100:
                    break
                row_count += 1
                for cell in row:
                    if cell and self._looks_like_ioc(str(cell)):
                        ioc_count += 1

            sheets.append(ExcelSheetInfo(
                name=sheet.title,
                index=wb.index(sheet),
                row_count=row_count,
                ioc_count=ioc_count
            ))

        logger.info(f"Found {len(sheets)} sheets in Excel file")
        return sheets

    def parse_sheets(
        self,
        file_path: Path,
        sheet_names: List[str],
        skip_header: bool = True
    ) -> ExcelParseResult:
        """
        Parse specific sheets from Excel file.

        Args:
            file_path: Path to Excel file
            sheet_names: List of sheet names to process
            skip_header: Whether to skip first row if it looks like header

        Returns:
            ExcelParseResult with data from selected sheets

        Raises:
            FileError: If file cannot be read
            CSVError: If parsing fails
        """
        self.error_handler.validate_file(file_path)

        try:
            import openpyxl
        except ImportError:
            raise CSVError(
                "Excel support requires openpyxl. "
                "Install with: pip install openpyxl"
            )

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            raise CSVError(f"Failed to load Excel file: {e}")

        result = ExcelParseResult(
            file_path=file_path,
            selected_sheets=sheet_names,
            encoding='utf-8',
            delimiter=','
        )

        total_iocs = 0

        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                continue

            sheet = wb[sheet_name]
            values = []

            # Extract all values from sheet
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        cleaned = str(cell).strip()
                        # Filter out non-IoC values (comments, empty, N/A, etc.)
                        if cleaned and not self._should_skip_value(cleaned):
                            values.append(cleaned)

            # Store result for this sheet
            result.sheet_results[sheet_name] = values
            total_iocs += len(values)
            result.total_rows += len(values)

        result.total_iocs = total_iocs

        logger.info(
            f"Parsed {len(sheet_names)} sheets from Excel file: "
            f"{result.total_rows} total rows, {result.total_iocs} values extracted"
        )

        return result

    def parse_all_sheets(
        self,
        file_path: Path,
        skip_header: bool = True
    ) -> ExcelParseResult:
        """
        Parse ALL sheets from Excel file.

        Args:
            file_path: Path to Excel file
            skip_header: Whether to skip first row if it looks like header

        Returns:
            ExcelParseResult with data from ALL sheets

        Raises:
            FileError: If file cannot be read
            CSVError: If parsing fails
        """
        try:
            import openpyxl
        except ImportError:
            raise CSVError(
                "Excel support requires openpyxl. "
                "Install with: pip install openpyxl"
            )

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            raise CSVError(f"Failed to load Excel file: {e}")

        result = ExcelParseResult(
            file_path=file_path,
            selected_sheets=[sheet.title for sheet in wb.worksheets],
            encoding='utf-8',
            delimiter=','
        )

        total_iocs = 0

        for sheet in wb.worksheets:
            values = []

            # Extract all values from sheet
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        cleaned = str(cell).strip()
                        # Filter out non-IoC values (comments, empty, N/A, etc.)
                        if cleaned and not self._should_skip_value(cleaned):
                            values.append(cleaned)

            # Store result for this sheet
            result.sheet_results[sheet.title] = values
            total_iocs += len(values)
            result.total_rows += len(values)

        result.total_iocs = total_iocs

        logger.info(f"Parsed all {len(wb.worksheets)} sheets from Excel file")

        return result

    def _looks_like_ioc(self, value: str) -> bool:
        """
        Quick check if a value looks like an IoC.

        Args:
            value: The value to check

        Returns:
            True if the value resembles an IoC
        """
        value = value.strip().lower()

        # SHA256: 64 hex chars
        if len(value) == 64 and all(c in "0123456789abcdef" for c in value):
            return True

        # SHA1: 40 hex chars
        if len(value) == 40 and all(c in "0123456789abcdef" for c in value):
            return True

        # MD5: 32 hex chars
        if len(value) == 32 and all(c in "0123456789abcdef" for c in value):
            return True

        # IPv4: 4 octets with dots
        if "." in value and value.count(".") == 3:
            parts = value.split(".")
            if len(parts) == 4:
                try:
                    return all(0 <= int(p) <= 255 for p in parts if p.isdigit())
                except ValueError:
                    pass

        return False

    def _should_skip_value(self, value: str) -> bool:
        """
        Check if a value should be skipped during parsing.

        Filters out:
        - Comments (lines starting with #, //, ;, --)
        - Empty/whitespace-only values
        - Single-character values (unlikely to be valid IoCs)
        - Common non-IoC placeholders (N/A, NULL, unknown, dashes, etc.)

        This matches the filtering behavior of CSVParser and IoCDetector.

        Args:
            value: The value to check

        Returns:
            True if the value should be skipped
        """
        # Check for empty/whitespace
        if not value or not value.strip():
            return True

        cleaned = value.strip()

        # Skip single-character values (unlikely to be valid IoCs)
        if len(cleaned) <= 1:
            return True

        # Skip comment lines
        for marker in self.COMMENT_PATTERNS:
            if cleaned.startswith(marker):
                return True

        # Skip common non-IoC patterns
        if cleaned.upper() in [p.upper() for p in self.NON_IOC_PATTERNS]:
            return True

        return False

    def is_excel_available(self) -> bool:
        """
        Check if openpyxl is available.

        Returns:
            True if openpyxl can be imported
        """
        try:
            import openpyxl
            return True
        except ImportError:
            return False
